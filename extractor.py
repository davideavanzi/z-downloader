import os
import pdfplumber
import re
import openpyxl
import argparse
from colorama import init, Fore

init(autoreset=True)  # Inizializza colorama


def estrai_testo_da_pdf(pdf_file):
    """Extract text from a PDF file."""
    testo = ""
    with pdfplumber.open(pdf_file) as pdf:
        for pagina in pdf.pages:
            testo += pagina.extract_text() or ""
    return testo


def extract_payslip_data(input_dir, output_dir):
    """Extract net salary data from payslip PDFs and save to Excel."""
    
    pdf_files = [f for f in os.listdir(input_dir) if f.lower().endswith(".pdf") and f.startswith("Libro_unico")]
    
    if not pdf_files:
        print(f"{Fore.YELLOW}No payslip PDFs found in {input_dir}{Fore.RESET}")
        return
    
    # Pattern to find amount after "NETTO DEL MESE" (handles spaces replaced with 's' or other chars)
    netto_pattern = re.compile(r'NETTO.?DEL.?MESE\s*\n?\s*([\d.,]+)\s*€', re.IGNORECASE)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Dati"

    sheet.cell(row=1, column=1, value="DATA")
    sheet.cell(row=1, column=2, value="NETTO DEL MESE")

    row_number = 2
    data_list = []

    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_dir, pdf_file)
        match = re.search(r'(\d{2}-\d{4})', pdf_file)
        if match:
            data = match.group(1)
        else:
            data = "Data non trovata"
        
        testo_pdf = estrai_testo_da_pdf(pdf_path)
        
        match_euro = netto_pattern.search(testo_pdf)
        if match_euro:
            euro = match_euro.group(1).strip()
            sheet.cell(row=row_number, column=1, value=data)
            sheet.cell(row=row_number, column=2, value=euro)
            data_list.append({"DATA": data, "NETTO DEL MESE": euro})
            row_number += 1
            print(f"File: {Fore.GREEN}{pdf_file}{Fore.RESET}, Data: {Fore.CYAN}{data}{Fore.RESET}, Importo: {Fore.GREEN}{euro}{Fore.RESET}")
        else:
            print(f"File: {Fore.RED}{pdf_file}{Fore.RESET}, Data: {Fore.CYAN}{data}{Fore.RESET}, {Fore.RED}Nessun importo trovato{Fore.RESET}")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    output_file = os.path.join(output_dir, "output.xls")
    workbook.save(output_file)
    print(f"\n{Fore.GREEN}Data saved to {output_file}{Fore.RESET}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        prog="Payslip Extractor",
        description="Extract net salary data from Zucchetti payslip PDFs"
    )
    parser.add_argument("-i", "--input-dir", default=".", help="Directory containing payslip PDFs (default: current directory)")
    parser.add_argument("-o", "--output-dir", default=".", help="Output directory for Excel file (default: current directory)")
    
    args = parser.parse_args()
    
    extract_payslip_data(args.input_dir, args.output_dir)
